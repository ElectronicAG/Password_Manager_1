################################################
# Password_Manager By Alan Gomez
# Python 3.13.3
# Kernel 6.12.28-1-MANJARO (64-bit)
#"""
#===============================================================
#⚠️ IMPORTANT NOTICE - EDUCATIONAL PROJECT ⚠️

#This code is a personal project created for educational purposes
#as I learn Python and database management.

#It does NOT implement any real security mechanisms:
#- No encryption of passwords or sensitive data.
#- No protection against attacks, unauthorized access, or major bugs.
#- Not intended for production or real-world password storage.

#💡 ANYONE WHO USES THIS CODE DOES SO AT THEIR OWN RISK.
#I am NOT responsible for any data loss, security breach, or misuse.

#If you need a real password manager, please use professional software 
#Thanks for understanding that this is part of my journey as a developer.
#===============================================================
#"""
################################################
import tkinter as tk  # Importa el módulo tkinter para crear interfaces gráficas
from tkinter import ttk, messagebox,font # Importa ttk para widgets temáticos y messagebox para cuadros de mensaje
import os  # Importa el módulo os para interactuar con el sistema operativo
import openpyxl  # Importa openpyxl para trabajar con archivos Excel
import subprocess  # Importa subprocess para ejecutar comandos del sistema
import string  # Importa el módulo string para trabajar con caracteres
import random  # Importa el módulo random para generar contraseñas aleatorias
import pandas as pd
from openpyxl import Workbook, load_workbook

#messagebox.showinfo("Archivo Existente", f"El archivo {archivo} ya existe") PARA INFO

VENTANA_SIZE = "650x400" ##Tamaño de las ventanas

#Terminal_Buscar
WIDTH_BOX_ENTRY_ABRIR_TERMINAL_BUSCAR = "60"
PADX_BOX_ENTRY_ABRIR_TERMINAL_BUSCAR = 10
PADY_BOX_ENTRY_ABRIR_TERMINAL_BUSCAR = 5

# Ruta de la carpeta y archivo
#(((((((((((((((((((((((((((((((((((((((((((((((((((((((((((((((((((((((
#(((((((((((((((((((((((((((((((((((((((((((((((((((((((((((((((((((((((
carpeta = r""  # Define la ruta de la carpeta
#(((((((((((((((((((((((((((((((((((((((((((((((((((((((((((((((((((((((
#(((((((((((((((((((((((((((((((((((((((((((((((((((((((((((((((((((((((
archivo = os.path.join(carpeta, "Data_P.xlsx")  # Define la ruta completa del archivo Excel

# Verificar si la carpeta existe, si no, crearla
if not os.path.exists(carpeta):  # Comprueba si la carpeta no existe
    os.makedirs(carpeta)  # Crea la carpeta
    # Establecer la carpeta como oculta
    subprocess.call(["attrib", "+h", carpeta])
    # messagebox.showinfo("Carpeta Creada", f"Se ha creado y ocultado la carpeta en {carpeta}")  # Muestra un mensaje informando que la carpeta fue creada y oculta

# Verificar si el archivo Excel existe, si no, crearlo
if not os.path.exists(archivo):  # Comprueba si el archivo no existe
    wb = openpyxl.Workbook()  # Crea un nuevo libro de Excel
    ws = wb.active  # Obtiene la hoja activa

    # Agregar encabezados de columnas
    ws.append(["Name","Link","Email","Password" ])  # Añade las cabeceras a la primera fila

    wb.save(archivo)  # Guarda el libro en la ruta especificada
    # messagebox.showinfo("Archivo Creado", f"Se ha creado el archivo {archivo}")  # Muestra un mensaje informando que el archivo fue creado


def abrir_terminal_buscar():
    root.destroy()  # Cierra la ventana principal
    buscar_window = tk.Tk()  # Crea una nueva ventana
    buscar_window.geometry(VENTANA_SIZE)
    
    buscar_window.title("Password_Manager")  # Establece el título de la ventana
    buscar_label = tk.Label(buscar_window, text="Aquí se buscará la contraseña.")  # Crea una etiqueta en la nueva ventana
    buscar_label.pack(pady=20)  # Añade un margen vertical a la etiqueta
    
    #########################################################################################################################
    # Cargar el archivo Excel
    try:
        df = pd.read_excel(archivo)
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo abrir el archivo: {e}")
        return

    # Variables para controlar los resultados y la posición actual
    resultados = None
    index_actual = 0

    # Función para mostrar un resultado específico basado en el índice actual
    def mostrar_resultado():
        nonlocal index_actual, resultados
        if resultados is not None and not resultados.empty:
            resultado = resultados.iloc[index_actual]  # Tomamos el resultado en la posición actual
            entry_name.delete(0, tk.END)
            entry_name.insert(0, resultado['Name'])
            entry_email.delete(0, tk.END)
            entry_email.insert(0, resultado['Email'])
            entry_link.delete(0, tk.END)
            entry_link.insert(0, resultado['Link'])
            entry_password.delete(0, tk.END)
            entry_password.insert(0, resultado['Password'])
        else:
            entry_name.delete(0, tk.END)
            entry_email.delete(0, tk.END)
            entry_link.delete(0, tk.END)
            entry_password.delete(0, tk.END)

    # Función para filtrar y mostrar los resultados
    def filtrar_resultados(*args):
        nonlocal index_actual, resultados
        busqueda = entry_busqueda.get()
        # Verificar que la búsqueda esté en mayúsculas
        if busqueda and not busqueda.isupper():
            entry_busqueda.delete(0, tk.END)
            return
        
        # Filtrar el DataFrame
        resultados = df[df['Name'].str.contains(busqueda, case=False, na=False)]
        index_actual = 0  # Reiniciar el índice al filtrar nuevos resultados
        mostrar_resultado()  # Mostrar el primer resultado

    # Función para mover hacia arriba en los resultados
    def subir_resultado():
        nonlocal index_actual
        if resultados is not None and not resultados.empty:
            if index_actual > 0:
                index_actual -= 1
                mostrar_resultado()

    # Función para mover hacia abajo en los resultados
    def bajar_resultado():
        nonlocal index_actual
        if resultados is not None and not resultados.empty:
            if index_actual < len(resultados) - 1:
                index_actual += 1
                mostrar_resultado()

    # Caja de texto para la búsqueda
    label_busqueda = tk.Label(buscar_window, text="Ingrese palabra (mayúsculas):")
    label_busqueda.pack()

    entry_busqueda = tk.Entry(buscar_window)
    entry_busqueda.pack()
    entry_busqueda.bind("<KeyRelease>", filtrar_resultados)  # Actualizar resultados en tiempo real

    # Crear etiquetas y cajas de texto para mostrar los valores de Name, Email, Link, Password
    label_name = tk.Label(buscar_window, text="Name:")
    label_name.pack(anchor="w", padx=PADX_BOX_ENTRY_ABRIR_TERMINAL_BUSCAR, pady=PADY_BOX_ENTRY_ABRIR_TERMINAL_BUSCAR)  # Alinea a la izquierda
    entry_name = tk.Entry(buscar_window, width=WIDTH_BOX_ENTRY_ABRIR_TERMINAL_BUSCAR)
    entry_name.pack(anchor="w", padx=PADX_BOX_ENTRY_ABRIR_TERMINAL_BUSCAR, pady=PADY_BOX_ENTRY_ABRIR_TERMINAL_BUSCAR)  # Alinea a la izquierda

    label_email = tk.Label(buscar_window, text="Email:")
    label_email.pack(anchor="w", padx=PADX_BOX_ENTRY_ABRIR_TERMINAL_BUSCAR, pady=PADY_BOX_ENTRY_ABRIR_TERMINAL_BUSCAR)
    entry_email = tk.Entry(buscar_window, width=WIDTH_BOX_ENTRY_ABRIR_TERMINAL_BUSCAR)
    entry_email.pack(anchor="w", padx=PADX_BOX_ENTRY_ABRIR_TERMINAL_BUSCAR, pady=PADY_BOX_ENTRY_ABRIR_TERMINAL_BUSCAR)

    label_link = tk.Label(buscar_window, text="Link:")
    label_link.pack(anchor="w", padx=PADX_BOX_ENTRY_ABRIR_TERMINAL_BUSCAR, pady=PADY_BOX_ENTRY_ABRIR_TERMINAL_BUSCAR)
    entry_link = tk.Entry(buscar_window, width=WIDTH_BOX_ENTRY_ABRIR_TERMINAL_BUSCAR)
    entry_link.pack(anchor="w", padx=PADX_BOX_ENTRY_ABRIR_TERMINAL_BUSCAR, pady=PADY_BOX_ENTRY_ABRIR_TERMINAL_BUSCAR)

    label_password = tk.Label(buscar_window, text="Password:")
    label_password.pack(anchor="w", padx=PADX_BOX_ENTRY_ABRIR_TERMINAL_BUSCAR, pady=PADY_BOX_ENTRY_ABRIR_TERMINAL_BUSCAR)
    entry_password = tk.Entry(buscar_window, width=WIDTH_BOX_ENTRY_ABRIR_TERMINAL_BUSCAR)
    entry_password.pack(anchor="w", padx=PADX_BOX_ENTRY_ABRIR_TERMINAL_BUSCAR, pady=PADY_BOX_ENTRY_ABRIR_TERMINAL_BUSCAR)

    # Botones para navegar entre los resultados
    button_subir = tk.Button(buscar_window, text="Subir", command=subir_resultado)
    button_subir.place(relx=1.0, rely=1.0, anchor='se', x=-50, y=-175)  # Coloca el botón en la esquina inferior derecha con un margen
    
    button_bajar = tk.Button(buscar_window, text="Bajar", command=bajar_resultado)
    button_bajar.place(relx=1.0, rely=1.0, anchor='se', x=-50, y=-125)  # Coloca el botón en la esquina inferior derecha con un margen

    #########################################################################################################################
    
    def close_terminal_buscar():
        buscar_window.destroy()
        main_window()

    regresar_button = tk.Button(buscar_window, text="return", command=close_terminal_buscar)
    regresar_button.place(relx=1.0, rely=1.0, anchor='se', x=-10, y=-10)  # Coloca el botón en la esquina inferior derecha con un margen
    buscar_window.mainloop()


def generate_password():
    length = int(slider.get())  # Obtiene la longitud de la contraseña del slider y la convierte a entero
    characters = ""  # Inicializa una cadena vacía para los caracteres de la contraseña
    if include_numbers.get():  # Si la opción de incluir números está seleccionada
        characters += string.digits  # Añade dígitos a los caracteres posibles
    if include_uppercase.get():  # Si la opción de incluir mayúsculas está seleccionada
        characters += string.ascii_uppercase  # Añade letras mayúsculas a los caracteres posibles
    if include_special.get():  # Si la opción de incluir caracteres especiales está seleccionada
        characters += string.punctuation  # Añade caracteres especiales a los caracteres posibles
    characters += string.ascii_lowercase  # Siempre añade letras minúsculas a los caracteres posibles

    password = ''.join(random.choice(characters) for _ in range(length))  # Genera la contraseña seleccionando caracteres aleatorios
    password_var.set(password)  # Establece la contraseña generada en la variable de la entrada

def update_label(value):
    slider_label.config(text=f"Longitud de la contraseña: {int(float(value))}")  # Actualiza la etiqueta del slider con el valor actual
    

def abrir_terminal_generar():
    root.destroy()  # Cierra la ventana principal
    generar_window = tk.Tk()  # Crea una nueva ventana
    generar_window.geometry("400x550")
    generar_window.title("Generar Contraseña")  # Establece el título de la ventana

    # Variables
    global include_numbers, include_uppercase, include_special, password_var, slider, slider_label
    include_numbers = tk.BooleanVar()  # Variable booleana para incluir números
    include_uppercase = tk.BooleanVar()  # Variable booleana para incluir mayúsculas
    include_special = tk.BooleanVar()  # Variable booleana para incluir caracteres especiales
    password_var = tk.StringVar()  # Variable para almacenar la contraseña generada

    # Widgets para generar contraseña
    slider_label = ttk.Label(generar_window, text="Longitud de la contraseña: 8")  # Etiqueta para mostrar la longitud de la contraseña
    slider_label.pack(pady=5)  # Añade un margen vertical a la etiqueta
    slider = ttk.Scale(generar_window, from_=3, to_=50, orient='horizontal', command=update_label)  # Crea un slider para seleccionar la longitud de la contraseña
    slider.set(8)  # Establece el valor inicial del slider
    slider.pack(pady=5)  # Añade un margen vertical al slider

    ttk.Checkbutton(generar_window, text="Incluir números", variable=include_numbers).pack(pady=5)  # Checkbox para incluir números
    ttk.Checkbutton(generar_window, text="Incluir letras mayúsculas", variable=include_uppercase).pack(pady=5)  # Checkbox para incluir mayúsculas
    ttk.Checkbutton(generar_window, text="Incluir caracteres especiales", variable=include_special).pack(pady=5)  # Checkbox para incluir caracteres especiales

    ttk.Button(generar_window, text="Generar Contraseña", command=generate_password).pack(pady=20)  # Botón para generar la contraseña
    ttk.Entry(generar_window, textvariable=password_var, width=50).pack(pady=5)  # Entrada para mostrar la contraseña generada

    # Entradas para agregar nuevo usuario
    name_var = tk.StringVar()  # Variable para el nombre
    link_var = tk.StringVar()  # Variable para el enlace
    email_var = tk.StringVar()  # Variable para el correo electrónico

    ttk.Label(generar_window, text="Nombre (mayúsculas)").pack(pady=5)
    ttk.Entry(generar_window, textvariable=name_var, width=50).pack(pady=5)

    ttk.Label(generar_window, text="Enlace").pack(pady=5)
    ttk.Entry(generar_window, textvariable=link_var, width=50).pack(pady=5)

    ttk.Label(generar_window, text="Correo Electrónico").pack(pady=5)
    ttk.Entry(generar_window, textvariable=email_var, width=50).pack(pady=5)

    def save_user_data():
        name = name_var.get()
        link = link_var.get()
        email = email_var.get()
        password = password_var.get()

        # Verificar que el nombre esté en mayúsculas
        if not name.isupper():
            messagebox.showerror("Error", "El nombre debe estar en mayúsculas.")
            #messagebox.showinfo("Archivo Existente", f"El archivo {archivo} ya existe") PARA INFO
            return  # Sale de la función si el nombre no es válido

        # Guardar los datos en el archivo de Excel
        wb = load_workbook(archivo)
        ws = wb.active
        ws.append([name, link, email, password])  # Agregar una nueva fila
        wb.save(archivo)

        # Limpiar las entradas
        name_var.set("")
        link_var.set("")
        email_var.set("")
        password_var.set("")

    ttk.Button(generar_window, text="Agregar Usuario", command=save_user_data).pack(pady=20)  # Botón para agregar nuevo usuario

    def close_terminal_generar():
        generar_window.destroy()
        main_window()

    regresar_button = tk.Button(generar_window, text="Regresar", command=close_terminal_generar)
    regresar_button.place(relx=1.0, rely=1.0, anchor='se', x=-10, y=-10)  # Coloca el botón en la esquina inferior derecha con un margen
    
    generar_window.mainloop()

def main_window():
    global root
    root = tk.Tk()  # Crea la ventana principal
    root.geometry(VENTANA_SIZE)
    root.title("Terminal de Contraseñas")  # Establece el título de la ventana principal
    
    # Crear una fuente personalizada
    custom_font = font.Font(family="Helvetica", size=14, weight="bold")  # Puedes cambiar el tipo y tamaño

    # Crear el botón con color, tamaño y fuente personalizados
    buscar_button = tk.Button(root, 
                              text="Password Manager", 
                              command=abrir_terminal_buscar, 
                              bg="#A020F0",       
                              fg="white",      
                              width=20,        
                              height=10,        
                              font=custom_font)  # Aplicar la fuente personalizada

    buscar_button.place(relx=1.0, rely=1.0, anchor='se', x=-225, y=0)

    generar_button = tk.Button(root, 
                           text="Generate Password",
                           command=abrir_terminal_generar, 
                           bg="#87CEEB",      # Color de fondo
                           fg="black",     # Color del textohu
                           width=20,       # Ancho del botón
                           height=10,
                           font=custom_font)       # Altura del botón


    generar_button.place(relx=1.0, rely=1.0, anchor='se', x=-0, y=0)
    
    
    root.geometry("450x210")  # Establece el tamaño de la ventana principal
    root.mainloop()  # Inicia el bucle principal de la ventana principal

main_window()
